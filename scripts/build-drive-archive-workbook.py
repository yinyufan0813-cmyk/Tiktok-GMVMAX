#!/usr/bin/env python3
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
LOGS = ROOT / "logs"
OUT = LOGS / "gmvmax-drive-archive-latest.xlsx"


def parse_num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", "")
    out = ""
    seen_dot = False
    for ch in text:
        if ch.isdigit() or ch == "-":
            out += ch
        elif ch == "." and not seen_dot:
            out += ch
            seen_dot = True
        elif out:
            break
    try:
        return float(out) if out not in ("", "-") else None
    except ValueError:
        return None


def read_jsonl(path):
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


def normalize_name(name):
    aliases = {
        "live-plan-1": "LIVE GMV Max_Gross revenue_YOUMILIER_20260529215644",
        "live-plan-2": "LIVE GMV Max_Gross revenue_YOUMILIER_20260521173451",
        "live-plan-3": "LIVE GMV Max_Gross revenue_YOUMILIER_20260519101516",
    }
    return aliases.get(name, name)


def write_table(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for idx, column in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in column:
            max_len = max(max_len, min(80, len(str(cell.value or ""))))
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    return ws


def decision_rows(snapshots):
    rows = []
    for snap in snapshots:
        for c in snap.get("campaigns", []):
            rows.append([
                snap.get("timestamp"),
                c.get("account"),
                normalize_name(c.get("name")),
                parse_num(c.get("totalSpend")),
                parse_num(c.get("totalOrderAmount")),
                parse_num(c.get("totalBudget")),
                parse_num(c.get("netSpend")),
                parse_num(c.get("intervalSpendIncrease")),
                parse_num(c.get("intervalOrderAmountIncrease")),
                parse_num(snap.get("totals", {}).get("roi")),
                snap.get("url"),
            ])
    return rows


def page_rows(pages):
    rows = []
    for snap in pages:
        for r in snap.get("tableState", {}).get("rows", []):
            cells = r.get("cells", [])
            if not cells or cells[0] in ("开/关", "广告计划名称"):
                continue
            account = cells[7].split(" ID:")[0] if len(cells) > 7 else ""
            rows.append([
                snap.get("timestamp") or snap.get("collectorTimestamp"),
                normalize_name(cells[0] if len(cells) > 0 else ""),
                cells[1] if len(cells) > 1 else "",
                parse_num(cells[2] if len(cells) > 2 else None),
                parse_num(cells[3] if len(cells) > 3 else None),
                parse_num(cells[4] if len(cells) > 4 else None),
                parse_num(cells[5] if len(cells) > 5 else None),
                parse_num(cells[6] if len(cells) > 6 else None),
                account,
                cells[8] if len(cells) > 8 else "",
                cells[10] if len(cells) > 10 else "",
            ])
    return rows


def network_summary(exchanges):
    by_family = Counter()
    by_key = Counter()
    lane_hits = Counter()
    lanes = ["bid", "rank", "candidate", "impression", "creative", "material", "recommendation", "delivery", "target_roi"]
    for row in exchanges:
        family = row.get("endpointFamily") or "unknown"
        key = row.get("endpointKey") or family
        by_family[family] += 1
        by_key[key] += 1
        text = " ".join(
            [family, key]
            + (row.get("requestBodyKeyPaths") or [])
            + (row.get("responseBodyKeyPaths") or [])
        ).lower()
        for lane in lanes:
            if lane == "target_roi":
                if "targetroi" in text or "target_roi" in text or "target roi" in text or "roas" in text:
                    lane_hits[lane] += 1
            elif lane in text:
                lane_hits[lane] += 1
    family_rows = [[k, v] for k, v in by_family.most_common(100)]
    key_rows = [[k, v] for k, v in by_key.most_common(200)]
    lane_rows = [[k, lane_hits[k]] for k in lanes]
    return family_rows, key_rows, lane_rows


def redacted_network_rows(exchanges):
    rows = []
    for row in exchanges:
        rows.append([
            row.get("timestampFinished") or row.get("collectorTimestamp") or row.get("timestamp"),
            row.get("endpointFamily") or "unknown",
            row.get("endpointKey") or "",
            row.get("method") or row.get("requestMethod") or "",
            row.get("status"),
            row.get("resourceType") or "",
            row.get("urlPath") or row.get("path") or safe_url_path(row.get("url") or row.get("requestUrl") or ""),
            ", ".join((row.get("requestBodyKeyPaths") or [])[:40]),
            ", ".join((row.get("responseBodyKeyPaths") or [])[:60]),
        ])
    return rows


def safe_url_path(url):
    if not url:
        return ""
    text = str(url)
    if "://" in text:
        text = text.split("://", 1)[1]
        text = text.split("/", 1)[1] if "/" in text else ""
    text = text.split("?", 1)[0]
    return "/" + text.lstrip("/") if text else ""


def allocation_rows(path):
    if not path.exists():
        return [], [], []
    data = json.loads(path.read_text(encoding="utf-8"))
    mix = data.get("campaignMix", [])
    mix_rows = [[
        r.get("account"),
        r.get("rows"),
        r.get("spendDelta"),
        r.get("spendShare"),
        r.get("avgWindowShare"),
        r.get("avgTargetRoi"),
        r.get("avgVisibleRoi"),
        r.get("avgRoiGap"),
    ] for r in mix]
    corr = data.get("globalCorrelation", {})
    corr_rows = [[k, v.get("n"), v.get("r")] for k, v in corr.items()]
    model = data.get("oneVariableModels", {})
    model_rows = [[k, v.get("n"), v.get("r2"), v.get("slope")] for k, v in model.items()]
    return mix_rows, corr_rows, model_rows


def manifest_rows():
    rows = []
    for path in sorted(LOGS.glob("*")):
        if path.is_file():
            stat = path.stat()
            rows.append([
                path.name,
                stat.st_size,
                datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "raw_local_only" if path.suffix in (".jsonl", ".log", ".png") else "synced_summary_or_tab",
            ])
    return rows


def main():
    decisions = read_jsonl(LOGS / "gmvmax-decision-snapshots.jsonl")
    pages = read_jsonl(LOGS / "gmvmax-page-snapshots.jsonl")
    exchanges = read_jsonl(LOGS / "gmvmax-network-exchanges.jsonl")
    material_pages = read_jsonl(LOGS / "material-page-snapshots.jsonl")
    material_records = read_jsonl(LOGS / "material-records.jsonl")

    wb = Workbook()
    wb.remove(wb.active)

    generated_at = datetime.now(timezone.utc).isoformat()
    latest_decision = decisions[-1].get("timestamp") if decisions else None
    latest_page = (pages[-1].get("timestamp") or pages[-1].get("collectorTimestamp")) if pages else None
    write_table(wb, "Index", ["key", "value"], [
        ["archive_name", "GMVMAX Google Drive archive snapshot"],
        ["generated_at_utc", generated_at],
        ["latest_decision_timestamp", latest_decision],
        ["latest_page_timestamp", latest_page],
        ["decision_snapshot_count", len(decisions)],
        ["page_snapshot_count", len(pages)],
        ["network_exchange_count", len(exchanges)],
        ["material_page_snapshot_count", len(material_pages)],
        ["material_record_count", len(material_records)],
        ["raw_network_note", "Raw JSONL is summarized here; full raw upload requires Drive raw-file upload/API access."],
    ])

    write_table(wb, "Decision Snapshots", [
        "timestamp", "account", "plan_name", "total_spend_myr", "total_gmv_myr",
        "total_budget_myr", "net_spend_myr", "interval_spend_myr",
        "interval_gmv_myr", "dashboard_roi", "url"
    ], decision_rows(decisions))

    write_table(wb, "Page Snapshots", [
        "timestamp", "plan_name", "status", "budget_myr", "cost_myr", "gmv_myr",
        "target_roi", "visible_roi", "account", "benefit", "optimization_mode"
    ], page_rows(pages))

    family_rows, key_rows, lane_rows = network_summary(exchanges)
    write_table(wb, "Network Family Summary", ["endpoint_family", "count"], family_rows)
    write_table(wb, "Network Key Summary", ["endpoint_key", "count"], key_rows)
    write_table(wb, "Network Lane Hits", ["lane", "key_path_hit_count"], lane_rows)
    write_table(wb, "Network Events Redacted", [
        "timestamp", "endpoint_family", "endpoint_key", "method", "status",
        "resource_type", "url_path", "request_key_paths", "response_key_paths"
    ], redacted_network_rows(exchanges))

    mix_rows, corr_rows, model_rows = allocation_rows(LOGS / "gmvmax-allocation-driver-analysis.json")
    write_table(wb, "Allocation Campaign Mix", [
        "account", "rows", "spend_delta_myr", "spend_share", "avg_window_share",
        "avg_target_roi", "avg_visible_roi", "avg_roi_gap"
    ], mix_rows)
    write_table(wb, "Allocation Correlations", ["feature", "n", "correlation_r"], corr_rows)
    write_table(wb, "Allocation Models", ["feature", "n", "r2", "slope"], model_rows)

    write_table(wb, "Local Raw Manifest", ["file", "bytes", "modified_utc", "sync_status"], manifest_rows())

    material_rows = []
    for row in material_records:
        material_rows.append([
            row.get("timestamp"),
            row.get("materialId") or row.get("id") or row.get("name"),
            row.get("planName") or row.get("campaignName"),
            row.get("status"),
            json.dumps(row, ensure_ascii=False)[:1000],
        ])
    write_table(wb, "Material Records Sample", ["timestamp", "material_id", "plan", "status", "raw_sample"], material_rows[:500])

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
